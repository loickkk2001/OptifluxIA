from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import re

# Mapping des couleurs RGB vers les spécialités
# Format: (R, G, B) -> nom_spécialité
COLOR_TO_SPECIALITY = {
    # Jaune - Urologie (approximatif, à ajuster selon les valeurs exactes)
    (255, 255, 0): "urologie",
    (255, 255, 153): "urologie",
    (255, 242, 204): "urologie",  # Jaune clair
    (255, 230, 153): "urologie",  # Jaune moyen
    
    # Bleu clair - Viscérale
    (173, 216, 230): "viscérale",  # Light blue
    (176, 224, 230): "viscérale",  # Powder blue
    (135, 206, 250): "viscérale",  # Light sky blue
    (173, 216, 230): "viscérale",  # Light blue
    
    # Vert - Vasculaire
    (0, 255, 0): "vasculaire",  # Vert pur
    (144, 238, 144): "vasculaire",  # Light green
    (152, 251, 152): "vasculaire",  # Pale green
    (124, 252, 0): "vasculaire",  # Lawn green
    
    # Rose - Hépatho gastro
    (255, 192, 203): "Hépatho gastro",  # Pink
    (255, 182, 193): "Hépatho gastro",  # Light pink
    (255, 20, 147): "Hépatho gastro",  # Deep pink
    (255, 105, 180): "Hépatho gastro",  # Hot pink
}

# Mapping des jours français vers index (0 = Lundi)
DAY_MAPPING = {
    "LUNDI": 0,
    "MARDI": 1,
    "MERCREDI": 2,
    "JEUDI": 3,
    "VENDREDI": 4,
}

def rgb_to_speciality(rgb: Optional[str]) -> Optional[str]:
    """
    Convertit une couleur RGB (format hex ou tuple) en nom de spécialité.
    
    Args:
        rgb: Couleur au format "RRGGBB" (hex) ou tuple (R, G, B)
    
    Returns:
        Nom de la spécialité ou None si non trouvée
    """
    if not rgb:
        return None
    
    # Si c'est une string hex (format "RRGGBB" ou "FFRRGGBB")
    if isinstance(rgb, str):
        rgb = rgb.upper()
        # Enlever le préfixe FF si présent
        if len(rgb) == 8 and rgb.startswith("FF"):
            rgb = rgb[2:]
        
        if len(rgb) == 6:
            try:
                r = int(rgb[0:2], 16)
                g = int(rgb[2:4], 16)
                b = int(rgb[4:6], 16)
                rgb_tuple = (r, g, b)
            except ValueError:
                return None
        else:
            return None
    elif isinstance(rgb, tuple) and len(rgb) == 3:
        rgb_tuple = rgb
    else:
        return None
    
    # Recherche exacte
    if rgb_tuple in COLOR_TO_SPECIALITY:
        return COLOR_TO_SPECIALITY[rgb_tuple]
    
    # Recherche approximative (tolérance de 30 pour chaque composante)
    tolerance = 30
    for (r_ref, g_ref, b_ref), speciality in COLOR_TO_SPECIALITY.items():
        if (abs(rgb_tuple[0] - r_ref) <= tolerance and
            abs(rgb_tuple[1] - g_ref) <= tolerance and
            abs(rgb_tuple[2] - b_ref) <= tolerance):
            return speciality
    
    return None


def parse_planning_excel(file_content: bytes, 
                        poll_id: str,
                        staff_id: str,
                        week_start_date: Optional[str] = None,
                        room_mapping: Optional[Dict[str, str]] = None) -> List[Dict]:
    """
    Parse un fichier Excel de planning avec feuilles POLE (structure TRAME/POLE réelle).
    Parcourt toutes les feuilles dont le nom commence par "POLE" et extrait les BoxPlans.
    """
    workbook = load_workbook(BytesIO(file_content), data_only=True)
    results = []

    # Parcourir les feuilles POLE ; si aucune, fallback sur feuille active
    for sheet_name in workbook.sheetnames:
        if sheet_name.upper().startswith("POLE"):
            sheet = workbook[sheet_name]
            results.extend(parse_pole_sheet(sheet, poll_id, staff_id, week_start_date, room_mapping))

    if not results:
        sheet = workbook.active
        results = parse_pole_sheet(sheet, poll_id, staff_id, week_start_date, room_mapping)

    return results


def parse_pole_sheet(sheet, poll_id: str, staff_id: str, week_start_date: Optional[str], room_mapping: Optional[Dict[str, str]] = None) -> List[Dict]:
    """
    Parse une feuille POLE structurée (discipline, N°salle, MATIN/AM, jours L-V).
    """
    results: List[Dict] = []
    current_discipline: Optional[str] = None
    current_room: Optional[str] = None
    current_phone: Optional[str] = None

    # Déterminer la date de début de semaine (lundi)
    if week_start_date:
        try:
            week_start = datetime.strptime(week_start_date, "%Y-%m-%d")
        except ValueError:
            week_start = datetime.now()
            week_start -= timedelta(days=week_start.weekday())
    else:
        week_start = datetime.now()
        week_start -= timedelta(days=week_start.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)

    # Colonnes des jours (G-K -> 7-11)
    DAY_COLS = {7: 0, 8: 1, 9: 2, 10: 3, 11: 4}

    for row in range(1, sheet.max_row + 1):
        # Discipline en colonne A (1)
        discipline_cell = sheet.cell(row=row, column=1)
        if discipline_cell.value and isinstance(discipline_cell.value, str):
            discipline_val = discipline_cell.value.strip()
            if discipline_val and discipline_val.upper() not in ["DISCIPLINE", "N°SALLE"]:
                current_discipline = discipline_val

        # N°salle en colonnes B-C (2-3), possible fusion
        room_cell_b = sheet.cell(row=row, column=2)
        room_cell_c = sheet.cell(row=row, column=3)
        room_value = f"{room_cell_b.value or ''} {room_cell_c.value or ''}".strip()
        if room_value and room_value.upper() != "N°SALLE":
            current_room = room_value

        # Téléphone en colonne D (4)
        phone_cell = sheet.cell(row=row, column=4)
        if phone_cell.value:
            phone_str = str(phone_cell.value).strip()
            phone_str = ''.join(c for c in phone_str if c.isdigit())
            if phone_str and len(phone_str) >= 4:
                current_phone = phone_str

        # Slots en colonne E (5) si existant
        slot_cell = sheet.cell(row=row, column=5)
        slot_value = slot_cell.value

        # Période en colonne F (6)
        period_cell = sheet.cell(row=row, column=6)
        period_value = str(period_cell.value).upper() if period_cell.value else ""
        if "MATIN" in period_value:
            current_period = "MATIN"
            period_time = "08:00"
        elif "AM" in period_value or "APRÈS-MIDI" in period_value or "APRES-MIDI" in period_value:
            current_period = "AM"
            period_time = "14:00"
        else:
            continue  # pas une ligne de période exploitable

        if not current_room or not current_phone:
            continue

        # Identifier room_id via mapping si fourni
        room_id = None
        if room_mapping and current_phone in room_mapping:
            room_id = room_mapping[current_phone]
        else:
            room_id = current_phone  # fallback: utiliser le téléphone comme identifiant

        # Parcourir jours Lundi-Vendredi (colonnes G-K)
        for col_idx, day_index in DAY_COLS.items():
            day_cell = sheet.cell(row=row, column=col_idx)
            if not day_cell.value:
                continue

            doctor_text = str(day_cell.value).strip()
            doctor_name, specialty = parse_doctor_cell(doctor_text, current_discipline)

            if not doctor_name:
                continue

            reservation_date = week_start + timedelta(days=day_index)

            entry = {
                "staff_id": staff_id,
                "doctors_id": [],  # mapping nom -> id à faire côté service si nécessaire
                "poll": poll_id,
                "room": room_id,
                "period": period_time,
                "date": reservation_date.strftime("%Y-%m-%d"),
                "consultation_number": str(slot_value or "0"),
                "consultation_time": "30",
                "comment": specialty or current_discipline or "",
                "status": "Réservé"
            }
            results.append(entry)

    return results


def parse_doctor_cell(text: str, default_specialty: Optional[str] = None) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse une cellule contenant \"Dr [NOM] - [Spécialité]\" ou variantes.
    """
    if not text:
        return None, None

    txt = text.strip()
    if not txt.upper().startswith("DR "):
        return None, None

    # retirer "Dr "
    txt = txt[3:].strip()

    # Séparer nom et spécialité
    parts = re.split(r'\s*[-–]\s*', txt, maxsplit=1)
    doctor_name = parts[0].strip()
    doctor_name = re.sub(r'\s+\d+.*$', '', doctor_name).strip()

    specialty = None
    if len(parts) > 1 and parts[1].strip():
        specialty = parts[1].strip()
    elif default_specialty:
        specialty = default_specialty

    return (doctor_name if doctor_name else None, specialty)

